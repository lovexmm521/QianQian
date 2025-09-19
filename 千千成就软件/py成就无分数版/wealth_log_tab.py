import json
import os
from datetime import datetime
import re

# 此功能需要 openpyxl 库，请通过命令 "pip install openpyxl" 来安装
try:
    import openpyxl
except ImportError:
    openpyxl = None

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                             QTableWidget, QTableWidgetItem, QHeaderView,
                             QMessageBox, QDialog, QDateEdit, QLineEdit, QFileDialog, QStyledItemDelegate)
from PyQt6.QtGui import QColor
from PyQt6.QtCore import Qt, pyqtSignal, QDate

# --- 常量 ---
WEALTH_LOG_FILE = "wealth_log.json"


class LargerEditDelegate(QStyledItemDelegate):
    """
    自定义委托，用于在QTableWidget中创建一个更大的编辑器。
    """
    def createEditor(self, parent, option, index):
        # 检查项目是否可编辑
        if not (index.flags() & Qt.ItemFlag.ItemIsEditable):
            return None

        editor = QLineEdit(parent)
        # 确保编辑器字体清晰可读
        font = parent.font()
        # 设置一个舒适的编辑字体大小
        font.setPointSize(11)
        editor.setFont(font)
        # 在编辑器内部添加一些内边距
        editor.setStyleSheet("padding: 2px;")
        return editor

    def updateEditorGeometry(self, editor, option, index):
        # 让编辑器的高度增加一些，使其看起来更大
        rect = option.rect
        rect.setHeight(rect.height() + 6)
        # 垂直居中
        rect.moveTop(rect.top() - 3)
        editor.setGeometry(rect)


class WealthLogTab(QWidget):
    wealth_updated = pyqtSignal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.log_data = []
        self.level_config = []
        self._is_populating = False
        self.init_ui()
        self.load_log()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["日期", "当时等级", "当前财富值", "趋势", "说明"])

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # 趋势列自适应宽度
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(4, 400)

        # 应用自定义委托来增大编辑框
        delegate = LargerEditDelegate(self.table)
        self.table.setItemDelegate(delegate)

        self.table.itemChanged.connect(self.handle_item_changed)

        button_layout = QHBoxLayout()
        add_button = QPushButton("➕ 添加记录")
        remove_button = QPushButton("➖ 删除选中行")
        import_button = QPushButton("📥 导入")
        export_button = QPushButton("📤 导出")

        button_layout.addWidget(add_button)
        button_layout.addWidget(remove_button)
        button_layout.addWidget(import_button)
        button_layout.addWidget(export_button)
        button_layout.addStretch()

        layout.addLayout(button_layout)
        layout.addWidget(self.table)

        add_button.clicked.connect(self.add_row_dialog)
        remove_button.clicked.connect(self.remove_row)
        import_button.clicked.connect(self.import_data)
        export_button.clicked.connect(self.export_data)

    def handle_item_changed(self, item):
        if self._is_populating:
            return

        row = item.row()
        col = item.column()

        # 从第0列获取存储的原始数据索引
        index_item = self.table.item(row, 0)
        if not index_item: return
        original_index = index_item.data(Qt.ItemDataRole.UserRole)

        if original_index is None or original_index >= len(self.log_data):
            return

        try:
            entry_to_update = self.log_data[original_index]

            if col == 0:  # 日期
                entry_to_update['date'] = item.text()
            elif col == 2:  # 财富值
                wealth_str = re.sub(r'[^\d]', '', item.text())
                entry_to_update['wealth'] = int(wealth_str)
            elif col == 4:  # 说明
                entry_to_update['description'] = item.text()
        except (ValueError, TypeError):
            self.refresh_table_and_emit_update()  # 如果输入格式错误，则刷新表格以撤销更改
            return

        self.save_log()

    def apply_settings(self, style_config):
        """应用显示设置，例如显示/隐藏趋势列"""
        show_trend = style_config.get("show_trend_column", False)
        self.table.setColumnHidden(3, not show_trend)

    def update_level_config(self, config_data):
        self.level_config = config_data

    def _get_level_data_for_wealth(self, wealth):
        """获取财富值对应的完整等级信息字典"""
        current_level_data = None
        sorted_config = sorted(self.level_config, key=lambda x: x['wealth_threshold'])
        for level in sorted_config:
            if wealth >= level['wealth_threshold']:
                current_level_data = level
            else:
                break
        return current_level_data

    def _get_level_for_wealth(self, wealth):
        level_data = self._get_level_data_for_wealth(wealth)
        if level_data:
            return level_data.get("level_name", f"等级 {level_data['level']}")
        return "未定级"

    def add_row_dialog(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("添加财富记录")
        layout = QVBoxLayout(dialog)
        date_edit = QDateEdit(QDate.currentDate())
        date_edit.setCalendarPopup(True)
        wealth_edit = QLineEdit()
        desc_edit = QLineEdit()

        form_layout = QHBoxLayout()
        form_layout.addWidget(QLabel("日期:"))
        form_layout.addWidget(date_edit)
        form_layout.addWidget(QLabel("当前财富值:"))
        form_layout.addWidget(wealth_edit)
        form_layout.addWidget(QLabel("说明:"))
        form_layout.addWidget(desc_edit)
        layout.addLayout(form_layout)

        button_box = QHBoxLayout()
        ok_button = QPushButton("确定")
        cancel_button = QPushButton("取消")
        button_box.addStretch()
        button_box.addWidget(ok_button)
        button_box.addWidget(cancel_button)
        layout.addLayout(button_box)

        ok_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)

        if dialog.exec():
            try:
                date = date_edit.date().toString("yyyy-MM-dd")
                current_wealth = int(wealth_edit.text())
                description = desc_edit.text()
                new_entry = {"date": date, "wealth": current_wealth, "description": description}
                self.log_data.append(new_entry)
                self.save_log()
            except ValueError:
                QMessageBox.warning(self, "输入错误", "当前财富值必须是一个整数。")

    def remove_row(self):
        current_row = self.table.currentRow()
        if current_row < 0: return

        index_item = self.table.item(current_row, 0)
        if not index_item: return
        original_index = index_item.data(Qt.ItemDataRole.UserRole)

        if original_index is not None and original_index < len(self.log_data):
            del self.log_data[original_index]
            self.save_log()

    def load_log(self):
        if os.path.exists(WEALTH_LOG_FILE):
            try:
                with open(WEALTH_LOG_FILE, 'r', encoding='utf-8') as f:
                    self.log_data = json.load(f)
            except json.JSONDecodeError:
                self.log_data = []
        self.refresh_table_and_emit_update()

    def _get_log_for_display(self):
        # 创建带原始索引的数据副本
        indexed_log = list(enumerate(self.log_data))
        try:
            # 根据日期排序
            sorted_indexed_log = sorted(
                indexed_log,
                key=lambda x: datetime.strptime(x[1].get('date', '1970-01-01'), '%Y-%m-%d'),
                reverse=True
            )
            return sorted_indexed_log
        except ValueError:
            return indexed_log  # 如果日期格式错误，返回原始顺序

    def refresh_table_and_emit_update(self):
        self._is_populating = True
        self.table.setRowCount(0)
        display_log_with_indices = self._get_log_for_display()

        for i, (original_index, entry) in enumerate(display_log_with_indices):
            row_count = self.table.rowCount()
            self.table.insertRow(row_count)

            # 日期 (可编辑), 并存储原始索引
            date_item = QTableWidgetItem(entry.get("date", ""))
            date_item.setData(Qt.ItemDataRole.UserRole, original_index)
            self.table.setItem(row_count, 0, date_item)

            # 当时等级 (不可编辑)
            level_name = self._get_level_for_wealth(entry.get('wealth', 0))
            level_item = QTableWidgetItem(level_name)
            level_item.setFlags(level_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_count, 1, level_item)

            # 当前财富值 (可编辑)
            current_wealth = entry.get('wealth', 0)
            wealth_item = QTableWidgetItem(f"{current_wealth:,}")
            wealth_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)  # 左对齐
            self.table.setItem(row_count, 2, wealth_item)

            # 趋势 (不可编辑)
            trend_item = QTableWidgetItem("")
            trend_item.setFlags(trend_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            trend_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            prev_original_index, prev_entry = display_log_with_indices[i + 1] if i + 1 < len(
                display_log_with_indices) else (None, None)
            if prev_entry:
                prev_wealth = prev_entry.get('wealth', 0)
                if current_wealth > prev_wealth:
                    trend_item.setText("↑")
                    trend_item.setForeground(QColor("#D32F2F"))
                elif current_wealth < prev_wealth:
                    trend_item.setText("↓")
                    trend_item.setForeground(QColor("#388E3C"))
            self.table.setItem(row_count, 3, trend_item)

            # 说明 (可编辑)
            self.table.setItem(row_count, 4, QTableWidgetItem(entry.get("description", "")))

        self._is_populating = False
        self.wealth_updated.emit(self.get_latest_wealth())

    def save_log(self):
        with open(WEALTH_LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.log_data, f, ensure_ascii=False, indent=4)
        self.refresh_table_and_emit_update()

    def get_latest_wealth(self):
        if not self.log_data: return 0
        try:
            latest_entry = max(self.log_data, key=lambda e: datetime.strptime(e.get('date', '1970-01-01'), '%Y-%m-%d'))
            return latest_entry.get('wealth', 0)
        except (ValueError, TypeError, KeyError):
            if self.log_data:
                return self.log_data[-1].get('wealth', 0)
            return 0

    def export_data(self):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导出功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "导出财富日志到 Excel", "", "Excel 文件 (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "财富日志"

            sheet["A1"] = "日期"
            sheet["B1"] = "当时等级"
            sheet["C1"] = "当前财富值"
            sheet["D1"] = "说明"

            # 获取带索引的排序后日志，但只使用其中的字典
            display_log = [item for _, item in self._get_log_for_display()]

            for row_index, entry in enumerate(display_log, start=2):
                sheet[f"A{row_index}"] = entry.get("date", "")
                level_name = self._get_level_for_wealth(entry.get('wealth', 0))
                sheet[f"B{row_index}"] = level_name
                sheet[f"C{row_index}"] = entry.get("wealth", 0)
                sheet[f"D{row_index}"] = entry.get("description", "")

            workbook.save(file_path)
            QMessageBox.information(self, "成功", f"财富日志已成功导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误: {e}")

    def import_data(self):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导入功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return

        file_path, _ = QFileDialog.getOpenFileName(self, "从 Excel 导入财富日志", "", "Excel 文件 (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            imported_logs = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 4: continue

                date_val, _, wealth_val, desc_val = row[0], row[1], row[2], row[3]

                if not date_val or wealth_val is None:
                    continue

                try:
                    if isinstance(date_val, datetime):
                        date_str = date_val.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date_val).split(" ")[0]

                    wealth = int(wealth_val)
                    description = str(desc_val) if desc_val is not None else ""

                    new_log = {
                        "date": date_str,
                        "wealth": wealth,
                        "description": description
                    }
                    imported_logs.append(new_log)
                except (ValueError, TypeError):
                    continue

            if not imported_logs:
                QMessageBox.warning(self, "导入提示", "在文件中没有找到有效的数据行。")
                return

            self.log_data.extend(imported_logs)
            self.save_log()
            QMessageBox.information(self, "成功", f"成功导入 {len(imported_logs)} 条财富记录。")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入过程中发生错误: {e}")

