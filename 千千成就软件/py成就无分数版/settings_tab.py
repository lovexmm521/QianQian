import os

# 此功能需要 openpyxl 库，请通过命令 "pip install openpyxl" 来安装
try:
    import openpyxl
except ImportError:
    openpyxl = None

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                             QTableWidget, QTableWidgetItem, QHeaderView, QFileDialog,
                             QMessageBox, QDialog, QGraphicsView, QGraphicsScene,
                             QGraphicsPixmapItem, QLineEdit, QStyledItemDelegate)
from PyQt6.QtGui import QPixmap, QTransform, QPainter
from PyQt6.QtCore import Qt, pyqtSignal, QSize


class LargerEditDelegate(QStyledItemDelegate):
    """
    自定义委托，用于在QTableWidget中创建一个更大的编辑器。
    """
    def createEditor(self, parent, option, index):
        # 检查项目是否可编辑
        if not (index.flags() & Qt.ItemFlag.ItemIsEditable):
            return None

        editor = QLineEdit(parent)
        # 确保编辑器字体清晰可读
        font = parent.font()
        # 设置一个舒适的编辑字体大小
        font.setPointSize(11)
        editor.setFont(font)
        # 在编辑器内部添加一些内边距
        editor.setStyleSheet("padding: 2px;")
        return editor

    def updateEditorGeometry(self, editor, option, index):
        # 让编辑器的高度增加一些，使其看起来更大
        rect = option.rect
        rect.setHeight(rect.height() + 6)
        # 垂直居中
        rect.moveTop(rect.top() - 3)
        editor.setGeometry(rect)


class ImagePreviewDialog(QDialog):
    def __init__(self, image_settings, parent=None, initial_size=None):
        super().__init__(parent)
        self.setWindowTitle("图片预览与设置")
        self.settings = image_settings.copy()
        self._initial_fit_done = False

        self.pixmap = QPixmap(self.settings.get("path", ""))
        if self.pixmap.isNull():
            QMessageBox.warning(self, "错误", "无法加载图片。")
            from PyQt6.QtCore import QTimer
            QTimer.singleShot(0, self.reject)
            return

        if self.settings.get("pos_x") is None or self.settings.get("pos_x") == 0:
            self.settings["pos_x"] = self.pixmap.width() / 2
        if self.settings.get("pos_y") is None or self.settings.get("pos_y") == 0:
            self.settings["pos_y"] = self.pixmap.height() / 2

        if initial_size:
            self.resize(initial_size)
        else:
            self.setMinimumSize(800, 600)

        self.init_ui()
        self.apply_settings()

    def init_ui(self):
        main_layout = QVBoxLayout(self)

        self.scene = QGraphicsScene(self)
        self.pixmap_item = QGraphicsPixmapItem(self.pixmap)
        self.scene.addItem(self.pixmap_item)

        self.view = QGraphicsView(self.scene)
        self.view.setDragMode(QGraphicsView.DragMode.ScrollHandDrag)
        self.view.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        self.view.setViewportUpdateMode(QGraphicsView.ViewportUpdateMode.FullViewportUpdate)
        self.view.setTransformationAnchor(QGraphicsView.ViewportAnchor.AnchorUnderMouse)
        self.view.setResizeAnchor(QGraphicsView.ViewportAnchor.AnchorViewCenter)

        controls_layout = QHBoxLayout()
        zoom_in_btn = QPushButton("放大 (+)")
        zoom_out_btn = QPushButton("缩小 (-)")
        rotate_btn = QPushButton("旋转 (↻)")
        reset_btn = QPushButton("重置视图")

        controls_layout.addWidget(zoom_in_btn)
        controls_layout.addWidget(zoom_out_btn)
        controls_layout.addWidget(rotate_btn)
        controls_layout.addStretch()
        controls_layout.addWidget(reset_btn)

        button_box = QHBoxLayout()
        ok_button = QPushButton("确定")
        cancel_button = QPushButton("取消")
        button_box.addStretch()
        button_box.addWidget(ok_button)
        button_box.addWidget(cancel_button)

        main_layout.addLayout(controls_layout)
        main_layout.addWidget(self.view)
        main_layout.addLayout(button_box)

        zoom_in_btn.clicked.connect(lambda: self.view.scale(1.2, 1.2))
        zoom_out_btn.clicked.connect(lambda: self.view.scale(1 / 1.2, 1 / 1.2))
        rotate_btn.clicked.connect(self.rotate_image)
        reset_btn.clicked.connect(self.reset_view)
        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)

    def showEvent(self, event):
        super().showEvent(event)
        if not self._initial_fit_done:
            is_default = (
                    abs(self.settings.get("zoom", 1.0) - 1.0) < 1e-6 and
                    self.settings.get("rotation", 0) == 0 and
                    abs(self.settings.get("pos_x", 0) - self.pixmap.width() / 2) < 1e-6 and
                    abs(self.settings.get("pos_y", 0) - self.pixmap.height() / 2) < 1e-6
            )
            if is_default:
                self.fit_image_in_view()
            self._initial_fit_done = True

    def fit_image_in_view(self):
        if self.pixmap.isNull():
            return
        self.view.fitInView(self.pixmap_item, Qt.AspectRatioMode.KeepAspectRatio)

    def apply_settings(self):
        self.pixmap_item.setTransformOriginPoint(self.pixmap.width() / 2, self.pixmap.height() / 2)
        self.pixmap_item.setRotation(self.settings.get("rotation", 0))

        transform = QTransform()
        zoom = self.settings.get("zoom", 1.0)
        transform.scale(zoom, zoom)
        self.view.setTransform(transform)

        pos_x = self.settings.get("pos_x", self.pixmap.width() / 2)
        pos_y = self.settings.get("pos_y", self.pixmap.height() / 2)
        self.view.centerOn(pos_x, pos_y)

    def rotate_image(self):
        current_rotation = self.pixmap_item.rotation()
        self.pixmap_item.setRotation(current_rotation + 90)

    def reset_view(self):
        self.pixmap_item.setRotation(0)
        self.fit_image_in_view()

    def accept(self):
        self.settings["rotation"] = self.pixmap_item.rotation() % 360
        self.settings["zoom"] = self.view.transform().m11()

        center_point = self.view.mapToScene(self.view.viewport().rect().center())
        self.settings["pos_x"] = center_point.x()
        self.settings["pos_y"] = center_point.y()

        super().accept()

    def get_settings(self):
        return self.settings


class SettingsTab(QWidget):
    config_updated = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.config_data = []
        self._is_populating = False
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["等级", "等级名称", "所需总财富", "奖励说明", "奖励图片路径"])

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)

        # 应用自定义委托来增大编辑框
        delegate = LargerEditDelegate(self.table)
        self.table.setItemDelegate(delegate)

        self.table.itemChanged.connect(self.handle_item_changed)

        button_layout = QHBoxLayout()
        add_button = QPushButton("➕ 添加一行")
        remove_button = QPushButton("➖ 删除选中行")
        import_button = QPushButton("📥 导入")
        export_button = QPushButton("📤 导出")
        self.settings_button = QPushButton("⚙️ 设置")

        button_layout.addWidget(add_button)
        button_layout.addWidget(remove_button)
        button_layout.addWidget(import_button)
        button_layout.addWidget(export_button)
        button_layout.addStretch()
        button_layout.addWidget(self.settings_button)

        layout.addLayout(button_layout)
        layout.addWidget(self.table)

        add_button.clicked.connect(self.add_row)
        remove_button.clicked.connect(self.remove_row)
        import_button.clicked.connect(self.import_data)
        export_button.clicked.connect(self.export_data)

    def handle_item_changed(self, item):
        if self._is_populating or item.column() == 0: return
        self.update_config_data_from_table();
        self.config_updated.emit()

    def load_from_data(self, data):
        self.config_data = data;
        self.populate_table()

    def add_row(self):
        self._is_populating = True;
        row_count = self.table.rowCount();
        self.table.insertRow(row_count)
        level_item = QTableWidgetItem(str(row_count + 1));
        level_item.setFlags(level_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table.setItem(row_count, 0, level_item);
        self.add_image_button(row_count);
        self._is_populating = False
        self.update_config_data_from_table()
        self.config_updated.emit()

    def remove_row(self):
        current_row = self.table.currentRow()
        if current_row >= 0: self.table.removeRow(
            current_row); self.update_config_data_from_table(); self.config_updated.emit()

    def add_image_button(self, row_index):
        cell_widget = QWidget()
        layout = QHBoxLayout(cell_widget)
        layout.setContentsMargins(2, 0, 2, 0)
        path_label = QLabel("未选择")
        browse_button = QPushButton("浏览...")
        browse_button.setObjectName("tableButton")
        customize_button = QPushButton("自定义")
        customize_button.setObjectName("tableButton")

        browse_button.clicked.connect(lambda _, r=row_index: self.browse_image(r))
        customize_button.clicked.connect(lambda _, r=row_index: self.customize_image(r))

        layout.addWidget(path_label, 1)
        layout.addWidget(browse_button)
        layout.addWidget(customize_button)
        self.table.setCellWidget(row_index, 4, cell_widget)

    def customize_image(self, row_index):
        cell_widget = self.table.cellWidget(row_index, 4)
        if not cell_widget: return

        image_settings = cell_widget.property("image_settings")
        if not image_settings or not image_settings.get("path") or not os.path.exists(image_settings.get("path")):
            QMessageBox.information(self, "提示", "请先为此行选择一个有效的图片。")
            return

        initial_size = None
        main_window = self.window()
        if main_window and hasattr(main_window, 'reward_image_display'):
            initial_size = main_window.reward_image_display.size()

        dialog = ImagePreviewDialog(image_settings, self, initial_size=initial_size)
        if dialog.exec():
            new_settings = dialog.get_settings()
            cell_widget.setProperty("image_settings", new_settings)
            self.update_config_data_from_table()
            self.config_updated.emit()

    def browse_image(self, row_index):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择奖励图片", "", "图片文件 (*.png *.jpg *.jpeg *.bmp)")
        if file_path:
            cell_widget = self.table.cellWidget(row_index, 4)
            path_label = cell_widget.findChild(QLabel)
            path_label.setText(os.path.basename(file_path));
            path_label.setToolTip(file_path)

            pixmap = QPixmap(file_path)
            new_settings = {
                "path": file_path, "zoom": 1.0, "rotation": 0,
                "pos_x": pixmap.width() / 2 if not pixmap.isNull() else 0,
                "pos_y": pixmap.height() / 2 if not pixmap.isNull() else 0
            }
            cell_widget.setProperty("image_settings", new_settings)

            self.update_config_data_from_table();
            self.config_updated.emit()

    def export_data(self):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导出功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "导出到 Excel", "", "Excel 文件 (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "等级设置"

            sheet["A1"] = "等级"
            sheet["B1"] = "等级名称"
            sheet["C1"] = "所需总财富"
            sheet["D1"] = "奖励说明"

            for row_index, level_data in enumerate(self.config_data, start=2):
                sheet[f"A{row_index}"] = level_data.get("level", "")
                sheet[f"B{row_index}"] = level_data.get("level_name", "")
                sheet[f"C{row_index}"] = level_data.get("wealth_threshold", 0)
                sheet[f"D{row_index}"] = level_data.get("reward_text", "")

            workbook.save(file_path)
            QMessageBox.information(self, "成功", f"数据已成功导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误: {e}")

    def import_data(self):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导入功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return

        file_path, _ = QFileDialog.getOpenFileName(self, "从 Excel 导入", "", "Excel 文件 (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            imported_levels = []
            # 从第二行开始读取
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 确保行中有足够的数据
                if len(row) < 4:
                    continue

                # B2: 等级名称, C2: 所需总财富, D2: 奖励说明
                level_name, wealth_threshold, reward_text = row[1], row[2], row[3]

                if not level_name or wealth_threshold is None:
                    continue  # 跳过等级名称或财富阈值为空的行

                try:
                    wealth = int(wealth_threshold)
                    new_level = {
                        "level_name": str(level_name),
                        "wealth_threshold": wealth,
                        "reward_text": str(reward_text) if reward_text is not None else "",
                        "reward_image": {"path": ""}
                    }
                    imported_levels.append(new_level)
                except (ValueError, TypeError):
                    # 如果财富值不是有效数字，则跳过此行
                    continue

            if not imported_levels:
                QMessageBox.warning(self, "导入提示", "在文件中没有找到有效的数据行。")
                return

            self.config_data.extend(imported_levels)

            # 重新排序和编号
            self.config_data.sort(key=lambda x: x['wealth_threshold'])
            for i, item in enumerate(self.config_data):
                item['level'] = i + 1

            self.populate_table()
            self.config_updated.emit()
            QMessageBox.information(self, "成功", f"成功导入 {len(imported_levels)} 条数据。")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入过程中发生错误: {e}")

    def populate_table(self):
        self._is_populating = True;
        self.table.setRowCount(0)
        for row_data in self.config_data:
            row_count = self.table.rowCount();
            self.table.insertRow(row_count)
            level_item = QTableWidgetItem(str(row_data.get("level", row_count + 1)));
            level_item.setFlags(level_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_count, 0, level_item)
            self.table.setItem(row_count, 1, QTableWidgetItem(row_data.get("level_name", "")))
            self.table.setItem(row_count, 2, QTableWidgetItem(str(row_data.get("wealth_threshold", 0))))
            self.table.setItem(row_count, 3, QTableWidgetItem(row_data.get("reward_text", "")))
            self.add_image_button(row_count)

            cell_widget = self.table.cellWidget(row_count, 4)
            image_settings = row_data.get("reward_image", {"path": ""})
            cell_widget.setProperty("image_settings", image_settings)

            path_label = cell_widget.findChild(QLabel)
            img_path = image_settings.get("path", "")
            if img_path: path_label.setText(os.path.basename(img_path)); path_label.setToolTip(img_path)
        self._is_populating = False

    def update_config_data_from_table(self):
        new_config_data = []
        for row in range(self.table.rowCount()):
            try:
                level_text = self.table.item(row, 0).text()
                name_item = self.table.item(row, 1)
                level_name = name_item.text() if name_item and name_item.text().strip() else f"等级 {level_text}"
                wealth_item = self.table.item(row, 2)
                if not wealth_item or not wealth_item.text():
                    default_settings = {"path": "", "zoom": 1.0, "rotation": 0, "pos_x": 0, "pos_y": 0}
                    cell_widget = self.table.cellWidget(row, 4)
                    if cell_widget and cell_widget.property("image_settings"):
                        default_settings = cell_widget.property("image_settings")

                    data = {"level": int(level_text), "level_name": level_name, "wealth_threshold": 0,
                            "reward_text": "", "reward_image": default_settings}
                    new_config_data.append(data)
                    continue

                wealth_threshold = int(wealth_item.text())
                reward_item = self.table.item(row, 3)
                reward_text = reward_item.text() if reward_item and reward_item.text() else ""

                cell_widget = self.table.cellWidget(row, 4)
                reward_image_settings = cell_widget.property("image_settings") if cell_widget else {"path": ""}

                data = {"level": int(level_text), "level_name": level_name, "wealth_threshold": wealth_threshold,
                        "reward_text": reward_text, "reward_image": reward_image_settings}
                new_config_data.append(data)
            except (ValueError, AttributeError):
                continue
        new_config_data.sort(key=lambda x: x['wealth_threshold'])
        for i, item in enumerate(new_config_data): item['level'] = i + 1
        self.config_data = new_config_data;
        self.populate_table()
